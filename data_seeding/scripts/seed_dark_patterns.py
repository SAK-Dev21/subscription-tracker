import openpyxl
import sys
import os

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))

from app import create_app
from app.extensions import db
from app.models import DarkPatternCategory, CognitiveBias, DarkPattern

app = create_app()

with app.app_context():
    # Clear existing data in reverse FK order
    db.session.query(DarkPattern).delete()
    db.session.query(DarkPatternCategory).delete()
    db.session.query(CognitiveBias).delete()
    db.session.commit()

    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__), '..', 'dark_patterns_data.xlsx'))

    # Seed categories
    cat_sheet = wb['Categories']
    cat_count = 0
    for row in cat_sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        db.session.add(DarkPatternCategory(name=row[0], description=row[1]))
        cat_count += 1
    db.session.commit()

    # Seed cognitive biases
    bias_sheet = wb['Cognitive Biases']
    bias_count = 0
    for row in bias_sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        db.session.add(CognitiveBias(name=row[0], description=row[1]))
        bias_count += 1
    db.session.commit()

    # Build lookup dicts for FK resolution
    categories = {c.name: c for c in DarkPatternCategory.query.all()}
    biases = {b.name: b for b in CognitiveBias.query.all()}

    # Seed dark patterns
    dp_sheet = wb['Dark Patterns']
    dp_count = 0
    for row in dp_sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        category = categories.get(row[1])
        if not category:
            print(f'Warning: category "{row[1]}" not found for pattern "{row[0]}" — skipping.')
            continue

        # Handle multiple biases — use first one only
        bias = None
        if row[6]:
            first_bias_name = str(row[6]).split(',')[0].strip()
            bias = biases.get(first_bias_name)

        db.session.add(DarkPattern(
            name=row[0],
            category_id=category.id,
            description=row[2],
            real_example=row[3],
            how_to_spot=row[4],
            how_to_protect=row[5],
            cognitive_bias_id=bias.id if bias else None,
            image_filename=row[7] if row[7] else None,
            source_citation=row[8],
        ))
        dp_count += 1
    db.session.commit()

    print(f'Loaded {cat_count} categories, {bias_count} cognitive biases, {dp_count} dark patterns.')
